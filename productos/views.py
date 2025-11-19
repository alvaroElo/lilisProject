from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.db.models import Q, Count, Sum, Avg
from django.core.paginator import Paginator
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .models import Producto
from maestros.models import Categoria, Marca, UnidadMedida


@login_required(login_url='login')
def productos_list(request):
    """Vista de listado de productos con filtros y paginación"""
    
    # Verificar permisos
    permisos = {
        'ver': True,
        'crear': True,
        'editar': True,
        'eliminar': True,
        'exportar': True
    }
    
    if hasattr(request.user, 'usuario_profile'):
        user_profile = request.user.usuario_profile
        if user_profile.rol and hasattr(user_profile.rol, 'permisos'):
            permisos_rol = user_profile.rol.permisos
            # Si permisos es un diccionario
            if isinstance(permisos_rol, dict):
                permisos_productos = permisos_rol.get('productos', {})
                if isinstance(permisos_productos, dict):
                    permisos['ver'] = permisos_productos.get('ver', True)
                    permisos['crear'] = permisos_productos.get('crear', True)
                    permisos['editar'] = permisos_productos.get('editar', True)
                    permisos['eliminar'] = permisos_productos.get('eliminar', True)
                    permisos['exportar'] = permisos_productos.get('exportar', True)
    
    if not permisos['ver']:
        return render(request, '403.html', status=403)
    
    # Obtener parámetros de búsqueda y filtros
    search = request.GET.get('search', '')
    categoria_filter = request.GET.get('categoria', '')
    marca_filter = request.GET.get('marca', '')
    estado_filter = request.GET.get('estado', '')
    bajo_stock = request.GET.get('bajo_stock', '')
    per_page = int(request.GET.get('per_page', 20))
    
    # Base queryset
    productos = Producto.objects.select_related('categoria', 'marca', 'uom_compra', 'uom_venta').all()
    
    # Aplicar filtros
    if search:
        productos = productos.filter(
            Q(sku__icontains=search) |
            Q(nombre__icontains=search) |
            Q(descripcion__icontains=search) |
            Q(ean_upc__icontains=search)
        )
    
    if categoria_filter:
        productos = productos.filter(categoria_id=categoria_filter)
    
    if marca_filter:
        productos = productos.filter(marca_id=marca_filter)
    
    if estado_filter:
        productos = productos.filter(estado=estado_filter)
    
    if bajo_stock == 'true':
        productos = productos.filter(alerta_bajo_stock=True)
    
    # Estadísticas
    total_productos = productos.count()
    productos_activos = productos.filter(estado='ACTIVO').count()
    productos_inactivos = productos.filter(estado='INACTIVO').count()
    productos_bajo_stock = productos.filter(alerta_bajo_stock=True).count()
    
    # Valor total del inventario
    valor_inventario = sum([
        (p.stock_actual * (p.costo_promedio or Decimal('0'))) 
        for p in productos if p.costo_promedio
    ])
    
    # Paginación
    paginator = Paginator(productos, per_page)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # Obtener categorías y marcas para filtros
    categorias = Categoria.objects.filter(activo=True).order_by('nombre')
    marcas = Marca.objects.filter(activo=True).order_by('nombre')
    unidades_medida = UnidadMedida.objects.filter(activo=True).order_by('codigo')
    
    context = {
        'productos': page_obj,
        'page_obj': page_obj,
        'total_productos': total_productos,
        'productos_activos': productos_activos,
        'productos_inactivos': productos_inactivos,
        'productos_bajo_stock': productos_bajo_stock,
        'valor_inventario': valor_inventario,
        'search': search,
        'categoria_filter': categoria_filter,
        'marca_filter': marca_filter,
        'estado_filter': estado_filter,
        'bajo_stock': bajo_stock,
        'per_page': per_page,
        'categorias': categorias,
        'marcas': marcas,
        'unidades_medida': unidades_medida,
        'permisos': permisos,
        'ESTADO_CHOICES': Producto.ESTADO_CHOICES,
    }
    
    return render(request, 'productos/productos_list.html', context)


@login_required(login_url='login')
def producto_create(request):
    """Vista para crear un nuevo producto (AJAX)"""
    
    # Verificar permisos
    if hasattr(request.user, 'usuario_profile'):
        user_profile = request.user.usuario_profile
        if user_profile.rol and user_profile.rol.permisos and isinstance(user_profile.rol.permisos, dict):
            permisos = user_profile.rol.permisos.get('productos', {})
            if not permisos.get('crear', True):
                return JsonResponse({
                    'success': False,
                    'message': 'No tienes permisos para crear productos'
                }, status=403)
    
    if request.method == 'POST':
        try:
            # Validar SKU único
            sku = request.POST.get('sku', '').strip()
            if Producto.objects.filter(sku=sku).exists():
                return JsonResponse({
                    'success': False,
                    'message': f'El SKU "{sku}" ya existe'
                }, status=400)
            
            # Validar EAN/UPC único si se proporciona
            ean_upc = request.POST.get('ean_upc', '').strip()
            if ean_upc and Producto.objects.filter(ean_upc=ean_upc).exists():
                return JsonResponse({
                    'success': False,
                    'message': f'El código EAN/UPC "{ean_upc}" ya existe'
                }, status=400)
            
            # Validar stock mínimo
            stock_minimo_value = request.POST.get('stock_minimo', '').strip()
            if not stock_minimo_value:
                return JsonResponse({
                    'success': False,
                    'message': 'El campo Stock Mínimo es obligatorio'
                }, status=400)
            
            try:
                stock_minimo = Decimal(stock_minimo_value)
                if stock_minimo < 0:
                    return JsonResponse({
                        'success': False,
                        'message': 'El Stock Mínimo no puede ser negativo'
                    }, status=400)
            except:
                return JsonResponse({
                    'success': False,
                    'message': 'El Stock Mínimo debe ser un número válido'
                }, status=400)
            
            # Crear producto
            producto = Producto.objects.create(
                sku=sku,
                ean_upc=ean_upc if ean_upc else None,
                nombre=request.POST.get('nombre'),
                descripcion=request.POST.get('descripcion', ''),
                categoria_id=request.POST.get('categoria'),
                marca_id=request.POST.get('marca') if request.POST.get('marca') else None,
                modelo=request.POST.get('modelo', ''),
                uom_compra_id=request.POST.get('uom_compra'),
                uom_venta_id=request.POST.get('uom_venta'),
                factor_conversion=Decimal(request.POST.get('factor_conversion', '1')),
                costo_estandar=Decimal(request.POST.get('costo_estandar', '0')) if request.POST.get('costo_estandar') else None,
                precio_venta=Decimal(request.POST.get('precio_venta', '0')) if request.POST.get('precio_venta') else None,
                impuesto_iva=Decimal(request.POST.get('impuesto_iva', '19')),
                stock_minimo=stock_minimo,
                stock_maximo=Decimal(request.POST.get('stock_maximo') or '0') if request.POST.get('stock_maximo') else None,
                punto_reorden=Decimal(request.POST.get('punto_reorden') or '0') if request.POST.get('punto_reorden') else None,
                perecible=request.POST.get('perecible') == 'on',
                control_por_lote=request.POST.get('control_por_lote') == 'on',
                control_por_serie=request.POST.get('control_por_serie') == 'on',
                imagen_url=request.POST.get('imagen_url', ''),
                ficha_tecnica_url=request.POST.get('ficha_tecnica_url', ''),
                estado=request.POST.get('estado', 'ACTIVO'),
            )
            
            return JsonResponse({
                'success': True,
                'message': f'Producto {producto.sku} creado exitosamente',
                'redirect': '/productos/'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error al crear producto: {str(e)}'
            }, status=500)
    
    return JsonResponse({'success': False, 'message': 'Método no permitido'}, status=405)


@login_required(login_url='login')
def producto_edit(request, producto_id):
    """Vista para editar un producto (AJAX)"""
    
    producto = get_object_or_404(Producto, id=producto_id)
    
    # Verificar permisos para edición
    if request.method == 'POST':
        if hasattr(request.user, 'usuario_profile'):
            user_profile = request.user.usuario_profile
            if user_profile.rol and user_profile.rol.permisos and isinstance(user_profile.rol.permisos, dict):
                permisos = user_profile.rol.permisos.get('productos', {})
                if not permisos.get('editar', True):
                    return JsonResponse({
                        'success': False,
                        'message': 'No tienes permisos para editar productos'
                    }, status=403)
    
    if request.method == 'POST':
        try:
            # Validar SKU único (excepto el actual)
            sku = request.POST.get('sku', '').strip()
            if Producto.objects.filter(sku=sku).exclude(id=producto_id).exists():
                return JsonResponse({
                    'success': False,
                    'message': f'El SKU "{sku}" ya existe en otro producto'
                }, status=400)
            
            # Validar EAN/UPC único si se proporciona
            ean_upc = request.POST.get('ean_upc', '').strip()
            if ean_upc and Producto.objects.filter(ean_upc=ean_upc).exclude(id=producto_id).exists():
                return JsonResponse({
                    'success': False,
                    'message': f'El código EAN/UPC "{ean_upc}" ya existe en otro producto'
                }, status=400)
            
            # Validar stock mínimo
            stock_minimo_value = request.POST.get('stock_minimo', '').strip()
            if not stock_minimo_value:
                return JsonResponse({
                    'success': False,
                    'message': 'El campo Stock Mínimo es obligatorio'
                }, status=400)
            
            try:
                stock_minimo = Decimal(stock_minimo_value)
                if stock_minimo < 0:
                    return JsonResponse({
                        'success': False,
                        'message': 'El Stock Mínimo no puede ser negativo'
                    }, status=400)
            except:
                return JsonResponse({
                    'success': False,
                    'message': 'El Stock Mínimo debe ser un número válido'
                }, status=400)
            
            # Actualizar datos
            producto.sku = sku
            producto.ean_upc = ean_upc if ean_upc else None
            producto.nombre = request.POST.get('nombre')
            producto.descripcion = request.POST.get('descripcion', '')
            producto.categoria_id = request.POST.get('categoria')
            producto.marca_id = request.POST.get('marca') if request.POST.get('marca') else None
            producto.modelo = request.POST.get('modelo', '')
            producto.uom_compra_id = request.POST.get('uom_compra')
            producto.uom_venta_id = request.POST.get('uom_venta')
            producto.factor_conversion = Decimal(request.POST.get('factor_conversion', '1'))
            producto.costo_estandar = Decimal(request.POST.get('costo_estandar', '0')) if request.POST.get('costo_estandar') else None
            producto.precio_venta = Decimal(request.POST.get('precio_venta', '0')) if request.POST.get('precio_venta') else None
            producto.impuesto_iva = Decimal(request.POST.get('impuesto_iva', '19'))
            producto.stock_minimo = stock_minimo
            producto.stock_maximo = Decimal(request.POST.get('stock_maximo') or '0') if request.POST.get('stock_maximo') else None
            producto.punto_reorden = Decimal(request.POST.get('punto_reorden') or '0') if request.POST.get('punto_reorden') else None
            producto.perecible = request.POST.get('perecible') == 'on'
            producto.control_por_lote = request.POST.get('control_por_lote') == 'on'
            producto.control_por_serie = request.POST.get('control_por_serie') == 'on'
            producto.imagen_url = request.POST.get('imagen_url', '')
            producto.ficha_tecnica_url = request.POST.get('ficha_tecnica_url', '')
            producto.estado = request.POST.get('estado', 'ACTIVO')
            
            producto.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Producto {producto.sku} actualizado exitosamente',
                'redirect': '/productos/'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error al actualizar producto: {str(e)}'
            }, status=500)
    
    # GET: devolver datos del producto
    return JsonResponse({
        'id': producto.id,
        'sku': producto.sku,
        'ean_upc': producto.ean_upc or '',
        'nombre': producto.nombre,
        'descripcion': producto.descripcion or '',
        'categoria': producto.categoria.id,
        'marca': producto.marca.id if producto.marca else '',
        'modelo': producto.modelo or '',
        'uom_compra': producto.uom_compra.id,
        'uom_venta': producto.uom_venta.id,
        'factor_conversion': str(producto.factor_conversion),
        'costo_estandar': str(producto.costo_estandar) if producto.costo_estandar else '',
        'costo_promedio': str(producto.costo_promedio) if producto.costo_promedio else '',
        'precio_venta': str(producto.precio_venta) if producto.precio_venta else '',
        'impuesto_iva': str(producto.impuesto_iva),
        'stock_actual': str(producto.stock_actual),
        'stock_minimo': str(producto.stock_minimo),
        'stock_maximo': str(producto.stock_maximo) if producto.stock_maximo else '',
        'punto_reorden': str(producto.punto_reorden) if producto.punto_reorden else '',
        'perecible': producto.perecible,
        'control_por_lote': producto.control_por_lote,
        'control_por_serie': producto.control_por_serie,
        'imagen_url': producto.imagen_url or '',
        'ficha_tecnica_url': producto.ficha_tecnica_url or '',
        'alerta_bajo_stock': producto.alerta_bajo_stock,
        'alerta_por_vencer': producto.alerta_por_vencer,
        'estado': producto.estado,
    })


@login_required(login_url='login')
def producto_delete(request, producto_id):
    """Vista para desactivar un producto (AJAX)"""
    
    # Verificar permisos
    if hasattr(request.user, 'usuario_profile'):
        user_profile = request.user.usuario_profile
        if user_profile.rol and user_profile.rol.permisos and isinstance(user_profile.rol.permisos, dict):
            permisos = user_profile.rol.permisos.get('productos', {})
            if not permisos.get('eliminar', True):
                return JsonResponse({
                    'success': False,
                    'message': 'No tienes permisos para desactivar productos'
                }, status=403)
    
    if request.method == 'POST':
        try:
            producto = get_object_or_404(Producto, id=producto_id)
            
            # Cambiar estado a INACTIVO en lugar de eliminar
            producto.estado = 'INACTIVO'
            producto.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Producto {producto.sku} desactivado exitosamente.'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error al desactivar producto: {str(e)}'
            }, status=500)
    
    return JsonResponse({'success': False, 'message': 'Método no permitido'}, status=405)


@login_required(login_url='login')
def exportar_productos_excel(request):
    """Exportar productos a Excel"""
    
    # Verificar permisos
    if hasattr(request.user, 'usuario_profile'):
        user_profile = request.user.usuario_profile
        if user_profile.rol and user_profile.rol.permisos and isinstance(user_profile.rol.permisos, dict):
            permisos = user_profile.rol.permisos.get('productos', {})
            if not permisos.get('exportar', True):
                return JsonResponse({
                    'success': False,
                    'message': 'No tienes permisos para exportar productos'
                }, status=403)
    
    # Obtener productos con los mismos filtros que la vista principal
    productos = Producto.objects.select_related('categoria', 'marca', 'uom_compra', 'uom_venta').all()
    
    # Aplicar filtros de búsqueda si existen
    search = request.GET.get('search', '').strip()
    if search:
        productos = productos.filter(
            Q(sku__icontains=search) |
            Q(nombre__icontains=search) |
            Q(descripcion__icontains=search) |
            Q(ean_upc__icontains=search)
        )
    
    categoria_filter = request.GET.get('categoria', '')
    if categoria_filter:
        productos = productos.filter(categoria_id=categoria_filter)
    
    marca_filter = request.GET.get('marca', '')
    if marca_filter:
        productos = productos.filter(marca_id=marca_filter)
    
    estado_filter = request.GET.get('estado', '')
    if estado_filter:
        productos = productos.filter(estado=estado_filter)
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"
    
    # Estilos
    header_fill = PatternFill(start_color="D20A11", end_color="D20A11", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = [
        'SKU', 'EAN/UPC', 'Nombre', 'Descripción', 'Categoría', 'Marca', 'Modelo',
        'UOM Compra', 'UOM Venta', 'Factor Conv.', 'Costo Estándar', 'Costo Promedio',
        'Precio Venta', 'IVA %', 'Stock Actual', 'Stock Mínimo', 'Stock Máximo',
        'Punto Reorden', 'Perecedero', 'Control Lote', 'Control Serie', 'Estado'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Datos
    for row_num, producto in enumerate(productos, 2):
        ws.cell(row=row_num, column=1).value = producto.sku
        ws.cell(row=row_num, column=2).value = producto.ean_upc or ''
        ws.cell(row=row_num, column=3).value = producto.nombre
        ws.cell(row=row_num, column=4).value = producto.descripcion or ''
        ws.cell(row=row_num, column=5).value = producto.categoria.nombre
        ws.cell(row=row_num, column=6).value = producto.marca.nombre if producto.marca else ''
        ws.cell(row=row_num, column=7).value = producto.modelo or ''
        ws.cell(row=row_num, column=8).value = producto.uom_compra.codigo
        ws.cell(row=row_num, column=9).value = producto.uom_venta.codigo
        ws.cell(row=row_num, column=10).value = float(producto.factor_conversion)
        ws.cell(row=row_num, column=11).value = float(producto.costo_estandar) if producto.costo_estandar else None
        ws.cell(row=row_num, column=12).value = float(producto.costo_promedio) if producto.costo_promedio else None
        ws.cell(row=row_num, column=13).value = float(producto.precio_venta) if producto.precio_venta else None
        ws.cell(row=row_num, column=14).value = float(producto.impuesto_iva)
        ws.cell(row=row_num, column=15).value = float(producto.stock_actual)
        ws.cell(row=row_num, column=16).value = float(producto.stock_minimo)
        ws.cell(row=row_num, column=17).value = float(producto.stock_maximo) if producto.stock_maximo else None
        ws.cell(row=row_num, column=18).value = float(producto.punto_reorden) if producto.punto_reorden else None
        ws.cell(row=row_num, column=19).value = 'Sí' if producto.perecible else 'No'
        ws.cell(row=row_num, column=20).value = 'Sí' if producto.control_por_lote else 'No'
        ws.cell(row=row_num, column=21).value = 'Sí' if producto.control_por_serie else 'No'
        ws.cell(row=row_num, column=22).value = producto.get_estado_display()
        
        # Aplicar bordes
        for col_num in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col_num).border = border
    
    # Ajustar ancho de columnas
    for col_num in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 15
    
    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=productos_dulceria_lilis.xlsx'
    wb.save(response)
    
    return response
