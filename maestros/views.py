from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import JsonResponse, HttpResponse
from .models import Proveedor
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime


@login_required(login_url='login')
def proveedores_list(request):
    """Vista principal del listado de proveedores"""
    
    # Obtener proveedores
    proveedores = Proveedor.objects.all()
    
    # Estadísticas generales (sin filtros)
    total_proveedores = Proveedor.objects.count()
    proveedores_activos = Proveedor.objects.filter(estado='ACTIVO').count()
    proveedores_bloqueados = Proveedor.objects.filter(estado='BLOQUEADO').count()
    proveedores_30_dias = Proveedor.objects.filter(condiciones_pago='30_DIAS').count()
    
    # Búsqueda
    search = request.GET.get('search', '').strip()
    if search:
        proveedores = proveedores.filter(
            Q(rut_nif__icontains=search) |
            Q(razon_social__icontains=search) |
            Q(nombre_fantasia__icontains=search) |
            Q(email__icontains=search) |
            Q(telefono__icontains=search) |
            Q(ciudad__icontains=search)
        )
    
    # Filtros
    estado_filter = request.GET.get('estado')
    if estado_filter:
        proveedores = proveedores.filter(estado=estado_filter)
    
    condicion_pago_filter = request.GET.get('condicion_pago')
    if condicion_pago_filter:
        proveedores = proveedores.filter(condiciones_pago=condicion_pago_filter)
    
    # Ordenamiento
    sort_by = request.GET.get('sort', 'created_at')
    sort_order = request.GET.get('order', 'desc')
    
    sort_fields = {
        'rut_nif': 'rut_nif',
        'razon_social': 'razon_social',
        'email': 'email',
        'telefono': 'telefono',
        'ciudad': 'ciudad',
        'condiciones_pago': 'condiciones_pago',
        'estado': 'estado',
        'created_at': 'created_at',
    }
    
    if sort_by in sort_fields:
        order_field = sort_fields[sort_by]
        if sort_order == 'desc':
            order_field = f'-{order_field}'
        proveedores = proveedores.order_by(order_field)
    
    # Paginación
    per_page = request.GET.get('per_page', 10)
    try:
        per_page = int(per_page)
    except:
        per_page = 10
    
    paginator = Paginator(proveedores, per_page)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'active_menu': 'proveedores',
        'proveedores': page_obj,
        'page_obj': page_obj,
        'search': search,
        'estado_filter': estado_filter,
        'condicion_pago_filter': condicion_pago_filter,
        'per_page': per_page,
        'sort_by': sort_by,
        'sort_order': sort_order,
        'ESTADO_CHOICES': Proveedor.ESTADO_CHOICES,
        'CONDICIONES_PAGO_CHOICES': Proveedor.CONDICIONES_PAGO_CHOICES,
        # Estadísticas
        'total_proveedores': total_proveedores,
        'proveedores_activos': proveedores_activos,
        'proveedores_bloqueados': proveedores_bloqueados,
        'proveedores_30_dias': proveedores_30_dias,
    }
    
    return render(request, 'proveedores/proveedores_list.html', context)


@login_required(login_url='login')
def proveedor_create(request):
    """Crear nuevo proveedor"""
    
    if request.method == 'POST':
        try:
            # Validar RUT único
            rut_nif = request.POST.get('rut_nif')
            if Proveedor.objects.filter(rut_nif=rut_nif).exists():
                return JsonResponse({
                    'success': False,
                    'message': f'El RUT/NIF "{rut_nif}" ya existe'
                })
            
            # Crear proveedor
            proveedor = Proveedor.objects.create(
                rut_nif=rut_nif,
                razon_social=request.POST.get('razon_social'),
                nombre_fantasia=request.POST.get('nombre_fantasia', '').strip() or None,
                email=request.POST.get('email'),
                telefono=request.POST.get('telefono', '').strip() or None,
                sitio_web=request.POST.get('sitio_web', '').strip() or None,
                direccion=request.POST.get('direccion', '').strip() or None,
                ciudad=request.POST.get('ciudad', '').strip() or None,
                pais=request.POST.get('pais', 'Chile'),
                condiciones_pago=request.POST.get('condiciones_pago'),
                condiciones_pago_detalle=request.POST.get('condiciones_pago_detalle', '').strip() or None,
                moneda=request.POST.get('moneda', 'CLP'),
                contacto_principal_nombre=request.POST.get('contacto_principal_nombre', '').strip() or None,
                contacto_principal_email=request.POST.get('contacto_principal_email', '').strip() or None,
                contacto_principal_telefono=request.POST.get('contacto_principal_telefono', '').strip() or None,
                estado=request.POST.get('estado', 'ACTIVO'),
                observaciones=request.POST.get('observaciones', '').strip() or None,
            )
            
            return JsonResponse({
                'success': True,
                'message': f'Proveedor "{proveedor.razon_social}" creado exitosamente'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error al crear proveedor: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'message': 'Método no permitido'})


@login_required(login_url='login')
def proveedor_edit(request, proveedor_id):
    """Editar proveedor existente"""
    
    proveedor = get_object_or_404(Proveedor, id=proveedor_id)
    
    if request.method == 'GET':
        # Retornar datos del proveedor
        return JsonResponse({
            'id': proveedor.id,
            'rut_nif': proveedor.rut_nif,
            'razon_social': proveedor.razon_social,
            'nombre_fantasia': proveedor.nombre_fantasia or '',
            'email': proveedor.email,
            'telefono': proveedor.telefono or '',
            'sitio_web': proveedor.sitio_web or '',
            'direccion': proveedor.direccion or '',
            'ciudad': proveedor.ciudad or '',
            'pais': proveedor.pais,
            'condiciones_pago': proveedor.condiciones_pago,
            'condiciones_pago_detalle': proveedor.condiciones_pago_detalle or '',
            'moneda': proveedor.moneda,
            'contacto_principal_nombre': proveedor.contacto_principal_nombre or '',
            'contacto_principal_email': proveedor.contacto_principal_email or '',
            'contacto_principal_telefono': proveedor.contacto_principal_telefono or '',
            'estado': proveedor.estado,
            'observaciones': proveedor.observaciones or '',
        })
    
    elif request.method == 'POST':
        try:
            # Validar RUT único (excepto el actual)
            rut_nif = request.POST.get('rut_nif')
            if Proveedor.objects.filter(rut_nif=rut_nif).exclude(id=proveedor_id).exists():
                return JsonResponse({
                    'success': False,
                    'message': f'El RUT/NIF "{rut_nif}" ya existe'
                })
            
            # Actualizar proveedor
            proveedor.rut_nif = rut_nif
            proveedor.razon_social = request.POST.get('razon_social')
            proveedor.nombre_fantasia = request.POST.get('nombre_fantasia', '').strip() or None
            proveedor.email = request.POST.get('email')
            proveedor.telefono = request.POST.get('telefono', '').strip() or None
            proveedor.sitio_web = request.POST.get('sitio_web', '').strip() or None
            proveedor.direccion = request.POST.get('direccion', '').strip() or None
            proveedor.ciudad = request.POST.get('ciudad', '').strip() or None
            proveedor.pais = request.POST.get('pais', 'Chile')
            proveedor.condiciones_pago = request.POST.get('condiciones_pago')
            proveedor.condiciones_pago_detalle = request.POST.get('condiciones_pago_detalle', '').strip() or None
            proveedor.moneda = request.POST.get('moneda', 'CLP')
            proveedor.contacto_principal_nombre = request.POST.get('contacto_principal_nombre', '').strip() or None
            proveedor.contacto_principal_email = request.POST.get('contacto_principal_email', '').strip() or None
            proveedor.contacto_principal_telefono = request.POST.get('contacto_principal_telefono', '').strip() or None
            proveedor.estado = request.POST.get('estado', 'ACTIVO')
            proveedor.observaciones = request.POST.get('observaciones', '').strip() or None
            proveedor.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Proveedor "{proveedor.razon_social}" actualizado exitosamente'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error al actualizar proveedor: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'message': 'Método no permitido'})


@login_required(login_url='login')
def proveedor_delete(request, proveedor_id):
    """Bloquear proveedor"""
    
    if request.method == 'POST':
        try:
            proveedor = get_object_or_404(Proveedor, id=proveedor_id)
            proveedor.estado = 'BLOQUEADO'
            proveedor.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Proveedor "{proveedor.razon_social}" bloqueado exitosamente'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error al bloquear proveedor: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'message': 'Método no permitido'})


@login_required(login_url='login')
def exportar_proveedores_excel(request):
    """Exportar listado de proveedores a Excel"""
    
    # Obtener proveedores con los mismos filtros
    proveedores = Proveedor.objects.all()
    
    # Aplicar filtros
    search = request.GET.get('search', '').strip()
    if search:
        proveedores = proveedores.filter(
            Q(rut_nif__icontains=search) |
            Q(razon_social__icontains=search) |
            Q(nombre_fantasia__icontains=search) |
            Q(email__icontains=search) |
            Q(telefono__icontains=search) |
            Q(ciudad__icontains=search)
        )
    
    estado_filter = request.GET.get('estado')
    if estado_filter:
        proveedores = proveedores.filter(estado=estado_filter)
    
    condicion_pago_filter = request.GET.get('condicion_pago')
    if condicion_pago_filter:
        proveedores = proveedores.filter(condiciones_pago=condicion_pago_filter)
    
    # Aplicar ordenamiento
    sort_by = request.GET.get('sort', 'created_at')
    sort_order = request.GET.get('order', 'desc')
    
    sort_fields = {
        'rut_nif': 'rut_nif',
        'razon_social': 'razon_social',
        'email': 'email',
        'telefono': 'telefono',
        'ciudad': 'ciudad',
        'condiciones_pago': 'condiciones_pago',
        'estado': 'estado',
        'created_at': 'created_at',
    }
    
    if sort_by in sort_fields:
        order_field = sort_fields[sort_by]
        if sort_order == 'desc':
            order_field = f'-{order_field}'
        proveedores = proveedores.order_by(order_field)
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Proveedores"
    
    # Estilos
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = ['RUT/NIF', 'Razón Social', 'Nombre Fantasía', 'Email', 'Teléfono', 
               'Ciudad', 'País', 'Condiciones Pago', 'Moneda', 'Estado', 'Fecha Creación']
    ws.append(headers)
    
    # Aplicar estilos a encabezados
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Ajustar ancho de columnas
    column_widths = [15, 35, 30, 30, 15, 20, 15, 18, 10, 12, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = width
    
    # Agregar datos
    for proveedor in proveedores:
        # Traducir condiciones de pago
        condiciones_dict = dict(Proveedor.CONDICIONES_PAGO_CHOICES)
        condiciones_text = condiciones_dict.get(proveedor.condiciones_pago, proveedor.condiciones_pago)
        
        row = [
            proveedor.rut_nif,
            proveedor.razon_social,
            proveedor.nombre_fantasia or '-',
            proveedor.email or '-',
            proveedor.telefono or '-',
            proveedor.ciudad or '-',
            proveedor.pais,
            condiciones_text,
            proveedor.moneda,
            'Activo' if proveedor.estado == 'ACTIVO' else 'Bloqueado',
            proveedor.created_at.strftime('%d/%m/%Y %H:%M') if proveedor.created_at else '-'
        ]
        ws.append(row)
        
        # Aplicar bordes y alineación
        for cell in ws[ws.max_row]:
            cell.border = border
            cell.alignment = Alignment(vertical='center')
    
    # Configurar respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    filename = f'proveedores_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    
    return response
