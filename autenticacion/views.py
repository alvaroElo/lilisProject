from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login as auth_login, logout as auth_logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import JsonResponse, HttpResponse
from productos.models import Producto
from inventario.models import AlertaStock, Bodega
from compras.models import OrdenCompra
from .models import Usuario, Rol
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime


def login_view(request):
    """Vista de login"""
    # Si ya está autenticado, redirigir al dashboard
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        # Autenticar usuario
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            # Login exitoso - limpiar contador de intentos fallidos
            if 'login_attempts' in request.session:
                del request.session['login_attempts']
            
            auth_login(request, user)
            
            # Redirigir a la página solicitada o al dashboard
            next_url = request.GET.get('next', 'dashboard')
            return redirect(next_url)
        else:
            # Login fallido - incrementar contador de intentos
            if 'login_attempts' not in request.session:
                request.session['login_attempts'] = {}
            
            attempts = request.session['login_attempts']
            
            # Incrementar intentos para este usuario
            if username:
                attempts[username] = attempts.get(username, 0) + 1
                request.session['login_attempts'] = attempts
                request.session.modified = True
                
                # Obtener número de intentos
                num_attempts = attempts[username]
                
                # Mensajes según número de intentos
                if num_attempts == 3:
                    messages.warning(request, f'⚠️ Has fallado 3 intentos de inicio de sesión con el usuario "{username}". Un error más y la cuenta será bloqueada.')
                elif num_attempts >= 4:
                    messages.error(request, f'🔒 La cuenta "{username}" ha sido bloqueada por múltiples intentos fallidos de inicio de sesión. (MODO PRUEBA - No se bloqueó realmente)')
                else:
                    messages.error(request, '🔒 Usuario o contraseña incorrectos. Por favor, verifica tus credenciales e intenta nuevamente.')
            else:
                messages.error(request, '🔒 Usuario o contraseña incorrectos. Por favor, verifica tus credenciales e intenta nuevamente.')
    
    return render(request, 'login/login.html')


@login_required(login_url='login')
def dashboard_view(request):
    """Vista del dashboard principal"""
    
    # Obtener estadísticas
    context = {
        'active_menu': 'dashboard',  # Para resaltar en el menú
        'productos_count': Producto.objects.filter(estado='ACTIVO').count(),
        'alertas_count': AlertaStock.objects.filter(estado='ACTIVA').count(),
        'ordenes_count': OrdenCompra.objects.filter(estado__in=['BORRADOR', 'ENVIADA', 'CONFIRMADA']).count(),
        'bodegas_count': Bodega.objects.filter(activo=True).count(),
        'notification_count': AlertaStock.objects.filter(estado='ACTIVA').count(),
    }
    
    return render(request, 'dashboard.html', context)


@login_required(login_url='login')
def logout_view(request):
    """Vista de logout"""
    auth_logout(request)
    
    # Limpiar toda la sesión
    request.session.flush()
    
    messages.info(request, 'Has cerrado sesión correctamente.')
    
    # Crear respuesta con headers anti-cache
    response = redirect('login')
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate, max-age=0'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    
    return response


@login_required(login_url='login')
def usuarios_list(request):
    """Vista de listado de usuarios con filtros y paginación"""
    
    # Verificar permisos de acceso al módulo
    if hasattr(request.user, 'usuario_profile'):
        usuario = request.user.usuario_profile
        if usuario.rol and usuario.rol.permisos and isinstance(usuario.rol.permisos, dict):
            permisos_usuarios = usuario.rol.permisos.get('usuarios', {})
            if not permisos_usuarios.get('ver', True):
                messages.error(request, 'No tienes permisos para acceder al módulo de usuarios')
                return redirect('dashboard')
    
    # Obtener parámetros de búsqueda y filtros
    search = request.GET.get('search', '')
    rol_filter = request.GET.get('rol', '')
    estado_filter = request.GET.get('estado', '')
    per_page = request.GET.get('per_page', '10')
    sort_by = request.GET.get('sort', 'created_at')
    sort_order = request.GET.get('order', 'desc')
    
    # Query base
    usuarios = Usuario.objects.select_related('user', 'rol').all()
    
    # Aplicar filtros
    if search:
        usuarios = usuarios.filter(
            Q(user__username__icontains=search) |
            Q(user__first_name__icontains=search) |
            Q(user__last_name__icontains=search) |
            Q(user__email__icontains=search)
        )
    
    if rol_filter:
        usuarios = usuarios.filter(rol__nombre=rol_filter)
    
    if estado_filter:
        usuarios = usuarios.filter(estado=estado_filter)
    
    # Mapeo de campos permitidos para ordenar
    sort_fields = {
        'username': 'user__username',
        'nombre': 'user__first_name',
        'email': 'user__email',
        'rol': 'rol__nombre',
        'estado': 'estado',
        'created_at': 'created_at',
    }
    
    # Validar y aplicar ordenamiento
    if sort_by in sort_fields:
        order_field = sort_fields[sort_by]
        if sort_order == 'asc':
            usuarios = usuarios.order_by(order_field)
        else:
            usuarios = usuarios.order_by(f'-{order_field}')
    else:
        # Ordenamiento por defecto
        usuarios = usuarios.order_by('-created_at')
    
    # Estadísticas
    total_usuarios = Usuario.objects.count()
    usuarios_activos = Usuario.objects.filter(estado='ACTIVO').count()
    usuarios_bloqueados = Usuario.objects.filter(estado='BLOQUEADO').count()
    usuarios_inactivos = Usuario.objects.filter(estado='INACTIVO').count()
    
    # Paginación
    try:
        per_page = int(per_page)
        if per_page not in [10, 20, 30]:
            per_page = 10
    except ValueError:
        per_page = 10
    
    paginator = Paginator(usuarios, per_page)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # Obtener todos los roles para el filtro
    roles = Rol.objects.all()
    
    # Obtener permisos del usuario actual
    permisos = {
        'ver': True,
        'crear': True,
        'editar': True,
        'eliminar': True,
        'exportar': True
    }
    
    if hasattr(request.user, 'usuario_profile'):
        usuario = request.user.usuario_profile
        if usuario.rol and usuario.rol.permisos and isinstance(usuario.rol.permisos, dict):
            rol_permisos = usuario.rol.permisos.get('usuarios', {})
            if rol_permisos:
                permisos = rol_permisos
    
    context = {
        'active_menu': 'usuarios',
        'page_obj': page_obj,
        'usuarios': page_obj.object_list,
        'total_usuarios': total_usuarios,
        'usuarios_activos': usuarios_activos,
        'usuarios_bloqueados': usuarios_bloqueados,
        'usuarios_inactivos': usuarios_inactivos,
        'roles': roles,
        'search': search,
        'rol_filter': rol_filter,
        'estado_filter': estado_filter,
        'per_page': per_page,
        'sort_by': sort_by,
        'sort_order': sort_order,
        'permisos': permisos,
    }
    
    return render(request, 'usuarios/usuarios_list.html', context)


@login_required(login_url='login')
def usuario_create(request):
    """Vista para crear un nuevo usuario (AJAX)"""
    
    # Verificar permisos
    if hasattr(request.user, 'usuario_profile'):
        usuario = request.user.usuario_profile
        if usuario.rol and usuario.rol.permisos and isinstance(usuario.rol.permisos, dict):
            permisos = usuario.rol.permisos.get('usuarios', {})
            if not permisos.get('crear', True):
                return JsonResponse({
                    'success': False,
                    'message': 'No tienes permisos para crear usuarios'
                }, status=403)
    
    if request.method == 'POST':
        try:
            # Datos del usuario Django
            username = request.POST.get('username')
            email = request.POST.get('email')
            first_name = request.POST.get('first_name')
            last_name = request.POST.get('last_name')
            password = request.POST.get('password')
            
            # Datos del perfil Usuario
            rol_id = request.POST.get('rol')
            telefono = request.POST.get('telefono', '')
            area_unidad = request.POST.get('area_unidad', '')
            estado = request.POST.get('estado', 'ACTIVO')
            foto_perfil = request.FILES.get('foto_perfil')
            
            # Validaciones básicas
            if not username or not email or not password or not rol_id:
                return JsonResponse({
                    'success': False,
                    'message': 'Todos los campos obligatorios deben ser completados.'
                }, status=400)
            
            # Verificar si el username ya existe
            if User.objects.filter(username=username).exists():
                return JsonResponse({
                    'success': False,
                    'message': f'El nombre de usuario "{username}" ya está en uso.'
                }, status=400)
            
            # Verificar si el email ya existe
            if User.objects.filter(email=email).exists():
                return JsonResponse({
                    'success': False,
                    'message': f'El email "{email}" ya está registrado.'
                }, status=400)
            
            # Crear usuario Django
            user = User.objects.create_user(
                username=username,
                email=email,
                first_name=first_name,
                last_name=last_name,
                password=password
            )
            
            # Crear perfil Usuario
            rol = Rol.objects.get(id=rol_id)
            usuario = Usuario.objects.create(
                user=user,
                rol=rol,
                telefono=telefono,
                area_unidad=area_unidad,
                estado=estado,
                foto_perfil=foto_perfil if foto_perfil else None
            )
            
            return JsonResponse({
                'success': True,
                'message': f'Usuario {username} creado exitosamente.',
                'redirect': '/usuarios/'
            })
            
        except Rol.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'El rol seleccionado no existe.'
            }, status=400)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error al crear usuario: {str(e)}'
            }, status=500)
    
    return JsonResponse({'success': False, 'message': 'Método no permitido.'}, status=405)


@login_required(login_url='login')
def usuario_edit(request, usuario_id):
    """Vista para editar un usuario (AJAX)"""
    
    usuario = get_object_or_404(Usuario, id=usuario_id)
    
    # No permitir editar usuarios administradores (excepto por otros administradores)
    if usuario.user.is_superuser or (usuario.rol and usuario.rol.nombre == 'ADMINISTRADOR'):
        if hasattr(request.user, 'usuario_profile'):
            user_profile = request.user.usuario_profile
            if not (request.user.is_superuser or (user_profile.rol and user_profile.rol.nombre == 'ADMINISTRADOR')):
                return JsonResponse({
                    'success': False,
                    'message': 'No tienes permisos para editar usuarios administradores.'
                }, status=403)
    
    # Verificar permisos para edición
    if request.method == 'POST':
        if hasattr(request.user, 'usuario_profile'):
            user_profile = request.user.usuario_profile
            if user_profile.rol and user_profile.rol.permisos and isinstance(user_profile.rol.permisos, dict):
                permisos = user_profile.rol.permisos.get('usuarios', {})
                if not permisos.get('editar', True):
                    return JsonResponse({
                        'success': False,
                        'message': 'No tienes permisos para editar usuarios'
                    }, status=403)
    
    if request.method == 'POST':
        try:
            # Actualizar datos del User de Django
            usuario.user.first_name = request.POST.get('first_name', '')
            usuario.user.last_name = request.POST.get('last_name', '')
            usuario.user.email = request.POST.get('email', '')
            
            # Actualizar contraseña solo si se proporciona
            new_password = request.POST.get('password', '')
            if new_password:
                usuario.user.set_password(new_password)
            
            usuario.user.save()
            
            # Actualizar datos del perfil Usuario
            rol_id = request.POST.get('rol')
            if rol_id:
                usuario.rol = Rol.objects.get(id=rol_id)
            
            usuario.telefono = request.POST.get('telefono', '')
            usuario.area_unidad = request.POST.get('area_unidad', '')
            usuario.estado = request.POST.get('estado', 'ACTIVO')
            
            # Actualizar foto de perfil si se proporciona
            foto_perfil = request.FILES.get('foto_perfil')
            if foto_perfil:
                usuario.foto_perfil = foto_perfil
            
            usuario.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Usuario {usuario.user.username} actualizado exitosamente.',
                'redirect': '/usuarios/'
            })
            
        except Rol.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'El rol seleccionado no existe.'
            }, status=400)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error al actualizar usuario: {str(e)}'
            }, status=500)
    
    # GET - Devolver datos del usuario en JSON
    return JsonResponse({
        'id': usuario.id,
        'username': usuario.user.username,
        'email': usuario.user.email,
        'first_name': usuario.user.first_name,
        'last_name': usuario.user.last_name,
        'rol_id': usuario.rol.id,
        'telefono': usuario.telefono or '',
        'area_unidad': usuario.area_unidad or '',
        'estado': usuario.estado,
        'foto_perfil_url': usuario.foto_perfil.url if usuario.foto_perfil else None,
    })


@login_required(login_url='login')
def usuario_delete(request, usuario_id):
    """Vista para desactivar un usuario (AJAX)"""
    
    # Verificar permisos
    if hasattr(request.user, 'usuario_profile'):
        user_profile = request.user.usuario_profile
        if user_profile.rol and user_profile.rol.permisos and isinstance(user_profile.rol.permisos, dict):
            permisos = user_profile.rol.permisos.get('usuarios', {})
            if not permisos.get('eliminar', True):
                return JsonResponse({
                    'success': False,
                    'message': 'No tienes permisos para desactivar usuarios'
                }, status=403)
    
    if request.method == 'POST':
        try:
            usuario = get_object_or_404(Usuario, id=usuario_id)
            
            # No permitir eliminar el propio usuario
            if usuario.user == request.user:
                return JsonResponse({
                    'success': False,
                    'message': 'No puedes desactivar tu propio usuario.'
                }, status=400)
            
            # No permitir eliminar usuarios superadmin o con rol ADMINISTRADOR
            if usuario.user.is_superuser or (usuario.rol and usuario.rol.nombre == 'ADMINISTRADOR'):
                return JsonResponse({
                    'success': False,
                    'message': 'No puedes desactivar usuarios administradores.'
                }, status=403)
            
            # Desactivar usuario (soft delete)
            usuario.estado = 'INACTIVO'
            usuario.user.is_active = False
            usuario.save()
            usuario.user.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Usuario {usuario.user.username} desactivado exitosamente.'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error al desactivar usuario: {str(e)}'
            }, status=500)
    
    return JsonResponse({'success': False, 'message': 'Método no permitido.'}, status=405)


def password_reset_request(request):
    """Vista para solicitar recuperación de contraseña"""
    
    if request.method == 'POST':
        email = request.POST.get('email')
        reset_link = None
        email_sent = False
        
        try:
            # Buscar usuario por email
            user = User.objects.get(email=email)
            
            # Generar token de recuperación
            from django.contrib.auth.tokens import default_token_generator
            from django.utils.http import urlsafe_base64_encode
            from django.utils.encoding import force_bytes
            
            token = default_token_generator.make_token(user)
            uid = urlsafe_base64_encode(force_bytes(user.pk))
            
            # Construir el link de recuperación
            reset_link = request.build_absolute_uri(
                f'/password-reset-confirm/{uid}/{token}/'
            )
            
            # Enviar email con Resend
            try:
                import resend
                from django.conf import settings
                
                resend.api_key = settings.RESEND_API_KEY
                
                # Nombre completo del usuario o username
                user_name = user.get_full_name() or user.username
                
                # HTML del email
                html_content = f"""
                <!DOCTYPE html>
                <html>
                <head>
                    <meta charset="UTF-8">
                    <style>
                        body {{ font-family: 'Segoe UI', Arial, sans-serif; line-height: 1.6; color: #333; }}
                        .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
                        .header {{ background: linear-gradient(135deg, #D20A11 0%, #8B0000 100%); color: white; padding: 30px; text-align: center; border-radius: 10px 10px 0 0; }}
                        .content {{ background: #f9f9f9; padding: 30px; border-radius: 0 0 10px 10px; }}
                        .button {{ display: inline-block; padding: 15px 30px; background: #D20A11; color: white; text-decoration: none; border-radius: 5px; font-weight: bold; margin: 20px 0; }}
                        .footer {{ text-align: center; margin-top: 30px; font-size: 12px; color: #666; }}
                        .warning {{ background: #fff3cd; border-left: 4px solid #ffc107; padding: 15px; margin: 20px 0; }}
                    </style>
                </head>
                <body>
                    <div class="container">
                        <div class="header">
                            <h1>🍬 {settings.COMPANY_NAME}</h1>
                            <p style="margin: 0;">Recuperación de Contraseña</p>
                        </div>
                        <div class="content">
                            <h2>Hola, {user_name}!</h2>
                            <p>Hemos recibido una solicitud para restablecer la contraseña de tu cuenta.</p>
                            <p>Haz clic en el siguiente botón para crear una nueva contraseña:</p>
                            
                            <div style="text-align: center;">
                                <a href="{reset_link}" class="button">Restablecer Contraseña</a>
                            </div>
                            
                            <p>O copia y pega este enlace en tu navegador:</p>
                            <p style="background: white; padding: 10px; border-radius: 5px; word-break: break-all; font-size: 12px;">
                                {reset_link}
                            </p>
                            
                            <div class="warning">
                                <strong>⚠️ Importante:</strong>
                                <ul style="margin: 10px 0;">
                                    <li>Este enlace es válido por <strong>24 horas</strong></li>
                                    <li>Si no solicitaste este cambio, ignora este correo</li>
                                    <li>Tu contraseña actual seguirá siendo válida</li>
                                </ul>
                            </div>
                            
                            <p style="margin-top: 30px; font-size: 14px; color: #666;">
                                Si tienes problemas con el botón, copia y pega el enlace directamente en tu navegador.
                            </p>
                        </div>
                        <div class="footer">
                            <p>Este correo fue enviado desde {settings.COMPANY_NAME}</p>
                            <p>© 2025 {settings.COMPANY_NAME}. Todos los derechos reservados.</p>
                        </div>
                    </div>
                </body>
                </html>
                """
                
                # En modo de prueba, enviar a email verificado
                # En producción con dominio verificado, enviar al email del usuario
                destination_email = settings.RESEND_TEST_EMAIL if (hasattr(settings, 'RESEND_TEST_EMAIL') and settings.RESEND_TEST_EMAIL) else email
                
                # Enviar email
                r = resend.Emails.send({
                    "from": settings.RESEND_FROM_EMAIL,
                    "to": destination_email,
                    "subject": f"🔐 Recuperación de Contraseña - {settings.COMPANY_NAME}",
                    "html": html_content
                })
                
                email_sent = True
                
                # Mensaje informativo en desarrollo
                if destination_email != email:
                    print(f"[DESARROLLO] Email de recuperación para {email} enviado a {destination_email}")
                
            except Exception as email_error:
                # Si falla el envío de email, registrar el error pero mostrar el link en pantalla
                import traceback
                error_detail = traceback.format_exc()
                print(f"Error al enviar email: {email_error}")
                print(f"Detalle completo: {error_detail}")
                # En ambiente de prueba, mostrar el link si falla el envío
                email_sent = False
            
        except User.DoesNotExist:
            # Por seguridad, no revelar que el email no existe
            # Mostrar el mismo mensaje genérico
            pass
        except Exception as e:
            messages.error(request, 'Error al procesar la solicitud. Intenta nuevamente.')
            return render(request, 'login/password_reset.html')
        
        # Siempre mostrar un mensaje genérico por seguridad
        if reset_link:
            if email_sent:
                # Email enviado exitosamente
                if hasattr(settings, 'RESEND_TEST_EMAIL') and settings.RESEND_TEST_EMAIL and email != settings.RESEND_TEST_EMAIL:
                    # Modo de desarrollo - informar que se envió al email de prueba
                    messages.success(request, f'✅ Correo enviado exitosamente a {settings.RESEND_TEST_EMAIL} (modo prueba)')
                    messages.info(request, f'ℹ️ En producción se enviaría a: {email}')
                else:
                    messages.success(request, f'Se ha enviado un correo a {email} con las instrucciones para restablecer tu contraseña.')
            else:
                # Falló el envío - mostrar el link (solo en desarrollo)
                messages.warning(request, 'El correo no pudo ser enviado. Usa el siguiente enlace:')
                return render(request, 'login/password_reset.html', {
                    'reset_link': reset_link,
                    'email': email
                })
        else:
            # Usuario no existe - mensaje genérico sin revelar
            messages.success(request, 'Si el correo está registrado, recibirás las instrucciones para restablecer tu contraseña.')
        
        return render(request, 'login/password_reset.html')
    
    return render(request, 'login/password_reset.html')


def password_reset_confirm(request, uidb64, token):
    """Vista para confirmar y cambiar la contraseña"""
    
    from django.contrib.auth.tokens import default_token_generator
    from django.utils.http import urlsafe_base64_decode
    from django.utils.encoding import force_str
    
    try:
        # Decodificar el uid
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = User.objects.get(pk=uid)
        
        # Verificar que el token sea válido
        if default_token_generator.check_token(user, token):
            
            if request.method == 'POST':
                new_password1 = request.POST.get('new_password1')
                new_password2 = request.POST.get('new_password2')
                
                # Validar que las contraseñas coincidan
                if new_password1 != new_password2:
                    messages.error(request, 'Las contraseñas no coinciden.')
                    return render(request, 'login/password_reset_confirm.html', {
                        'valid_link': True,
                        'email': user.email
                    })
                
                # Validar longitud mínima
                if len(new_password1) < 8:
                    messages.error(request, 'La contraseña debe tener al menos 8 caracteres.')
                    return render(request, 'login/password_reset_confirm.html', {
                        'valid_link': True,
                        'email': user.email
                    })
                
                # Cambiar la contraseña
                user.set_password(new_password1)
                user.save()
                
                messages.success(request, '¡Contraseña cambiada exitosamente! Ya puedes iniciar sesión.')
                return redirect('login')
            
            # GET - Mostrar formulario
            return render(request, 'login/password_reset_confirm.html', {
                'valid_link': True,
                'email': user.email
            })
        else:
            # Token inválido
            return render(request, 'login/password_reset_confirm.html', {
                'valid_link': False
            })
            
    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        return render(request, 'login/password_reset_confirm.html', {
            'valid_link': False
        })


@login_required(login_url='login')
def exportar_usuarios_excel(request):
    """Exportar listado de usuarios a Excel"""
    
    # Obtener usuarios con los mismos filtros que la vista principal
    usuarios = Usuario.objects.select_related('user', 'rol').all()
    
    # Aplicar filtros de búsqueda si existen
    search = request.GET.get('search', '').strip()
    if search:
        usuarios = usuarios.filter(
            Q(user__username__icontains=search) |
            Q(user__first_name__icontains=search) |
            Q(user__last_name__icontains=search) |
            Q(user__email__icontains=search) |
            Q(telefono__icontains=search)
        )
    
    # Filtros adicionales
    rol_filter = request.GET.get('rol')
    if rol_filter:
        usuarios = usuarios.filter(rol_id=rol_filter)
    
    estado_filter = request.GET.get('estado')
    if estado_filter:
        usuarios = usuarios.filter(estado=estado_filter)
    
    # Aplicar ordenamiento si existe
    sort_by = request.GET.get('sort', 'created_at')
    sort_order = request.GET.get('order', 'desc')
    
    sort_fields = {
        'username': 'user__username',
        'nombre': 'user__first_name',
        'email': 'user__email',
        'rol': 'rol__nombre',
        'estado': 'estado',
        'created_at': 'created_at',
    }
    
    if sort_by in sort_fields:
        order_field = sort_fields[sort_by]
        if sort_order == 'desc':
            order_field = f'-{order_field}'
        usuarios = usuarios.order_by(order_field)
    
    # Crear workbook de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Usuarios"
    
    # Estilos
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = ['Usuario', 'Nombre Completo', 'Email', 'Teléfono', 'Rol', 'Área/Unidad', 'Estado', 'Fecha Creación']
    ws.append(headers)
    
    # Aplicar estilos a encabezados
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Ajustar ancho de columnas
    column_widths = [15, 30, 35, 15, 20, 25, 12, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = width
    
    # Agregar datos
    for usuario in usuarios:
        row = [
            usuario.user.username,
            usuario.user.get_full_name() or '-',
            usuario.user.email or '-',
            usuario.telefono or '-',
            usuario.rol.nombre if usuario.rol else '-',
            usuario.area_unidad or '-',
            'Activo' if usuario.estado == 'ACTIVO' else 'Inactivo',
            usuario.created_at.strftime('%d/%m/%Y %H:%M') if usuario.created_at else '-'
        ]
        ws.append(row)
        
        # Aplicar bordes y alineación a las celdas de datos
        for cell in ws[ws.max_row]:
            cell.border = border
            cell.alignment = Alignment(vertical='center')
    
    # Configurar la respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Nombre del archivo con fecha y hora
    filename = f'usuarios_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    # Guardar el workbook en la respuesta
    wb.save(response)
    
    return response
