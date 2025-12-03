from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.db.models import Q, Count, Sum, F
from django.core.paginator import Paginator
from django.utils import timezone
from django.utils.timezone import localtime
from django.views.decorators.http import require_http_methods
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import json

from .models import MovimientoInventario, Bodega, Lote, StockActual
from productos.models import Producto
from maestros.models import Proveedor, UnidadMedida


def actualizar_stock_producto(movimiento):
    """
    Actualiza el stock del producto según el tipo de movimiento
    INGRESO/DEVOLUCION: Suma al stock
    SALIDA/AJUSTE: Resta del stock
    TRANSFERENCIA: Resta de origen, suma a destino
    """
    producto = movimiento.producto
    cantidad = movimiento.cantidad
    
    if movimiento.tipo_movimiento in ['INGRESO', 'DEVOLUCION']:
        # Aumentar stock
        producto.stock_actual += cantidad
    elif movimiento.tipo_movimiento in ['SALIDA', 'AJUSTE']:
        # Disminuir stock
        producto.stock_actual -= cantidad
    elif movimiento.tipo_movimiento == 'TRANSFERENCIA':
        # En transferencias, el stock total no cambia
        # Pero aquí simplemente ajustamos el stock general
        pass
    
    # Guardar producto (esto también actualiza alerta_bajo_stock automáticamente)
    producto.save()


@login_required
def movimientos_list(request):
    """Vista principal del listado de movimientos de inventario"""
    
    # Obtener permisos del usuario actual
    permisos = {
        'ver': True,
        'crear': True,
        'editar': True,
        'eliminar': True,
        'exportar': True
    }
    
    if hasattr(request.user, 'usuario_profile'):
        usuario = request.user.usuario_profile
        if usuario.rol and usuario.rol.permisos and isinstance(usuario.rol.permisos, dict):
            rol_permisos = usuario.rol.permisos.get('inventario', {})
            if rol_permisos:
                permisos.update(rol_permisos)  # Actualizar solo los permisos que existan en el rol
    
    if not permisos.get('ver'):
        return render(request, '403.html', status=403)
    
    # Obtener parámetros de filtros
    search = request.GET.get('search', '').strip()
    tipo_movimiento = request.GET.get('tipo_movimiento', '')
    estado = request.GET.get('estado', '')
    bodega_id = request.GET.get('bodega', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    per_page = request.GET.get('per_page', '25')
    sort_by = request.GET.get('sort', 'fecha_movimiento')
    sort_order = request.GET.get('order', 'desc')
    
    # Query base
    movimientos = MovimientoInventario.objects.select_related(
        'producto', 'bodega_origen', 'bodega_destino', 'proveedor', 
        'usuario__user', 'lote', 'unidad_medida'
    ).all()
    
    # Aplicar filtros
    if search:
        movimientos = movimientos.filter(
            Q(producto__sku__icontains=search) |
            Q(producto__nombre__icontains=search) |
            Q(documento_referencia__icontains=search) |
            Q(serie__icontains=search) |
            Q(lote__codigo_lote__icontains=search)
        )
    
    if tipo_movimiento:
        movimientos = movimientos.filter(tipo_movimiento=tipo_movimiento)
    
    if estado:
        movimientos = movimientos.filter(estado=estado)
    
    if bodega_id:
        movimientos = movimientos.filter(
            Q(bodega_origen_id=bodega_id) | Q(bodega_destino_id=bodega_id)
        )
    
    if fecha_desde:
        movimientos = movimientos.filter(fecha_movimiento__gte=fecha_desde)
    
    if fecha_hasta:
        movimientos = movimientos.filter(fecha_movimiento__lte=fecha_hasta + ' 23:59:59')
    
    # Mapeo de campos permitidos para ordenar
    sort_fields = {
        'fecha_movimiento': 'fecha_movimiento',
        'tipo_movimiento': 'tipo_movimiento',
        'producto': 'producto__nombre',
        'cantidad': 'cantidad',
        'estado': 'estado',
        'bodega_origen': 'bodega_origen__nombre',
        'bodega_destino': 'bodega_destino__nombre',
        'usuario': 'usuario__user__username',
        'created_at': 'created_at',
    }
    
    # Validar y aplicar ordenamiento
    if sort_by in sort_fields:
        order_field = sort_fields[sort_by]
        if sort_order == 'asc':
            movimientos = movimientos.order_by(order_field)
        else:
            movimientos = movimientos.order_by(f'-{order_field}')
    else:
        # Ordenamiento por defecto
        movimientos = movimientos.order_by('-fecha_movimiento', '-created_at')
    
    # Estadísticas (sin filtros para mostrar totales globales)
    total_movimientos = MovimientoInventario.objects.count()
    
    # Obtener fecha actual en timezone local y calcular rango del día
    now_local = localtime(timezone.now())
    inicio_dia = now_local.replace(hour=0, minute=0, second=0, microsecond=0)
    fin_dia = now_local.replace(hour=23, minute=59, second=59, microsecond=999999)
    
    movimientos_hoy = MovimientoInventario.objects.filter(
        fecha_movimiento__gte=inicio_dia,
        fecha_movimiento__lte=fin_dia
    ).count()
    movimientos_pendientes = MovimientoInventario.objects.filter(estado='PENDIENTE').count()
    
    # Calcular movimientos por tipo usando rango del mes actual en timezone local
    inicio_mes = now_local.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    # Calcular el último día del mes
    if now_local.month == 12:
        fin_mes = now_local.replace(year=now_local.year + 1, month=1, day=1, hour=0, minute=0, second=0, microsecond=0) - timedelta(microseconds=1)
    else:
        fin_mes = now_local.replace(month=now_local.month + 1, day=1, hour=0, minute=0, second=0, microsecond=0) - timedelta(microseconds=1)
    
    ingresos_mes = MovimientoInventario.objects.filter(
        tipo_movimiento='INGRESO',
        estado='CONFIRMADO',
        fecha_movimiento__gte=inicio_mes,
        fecha_movimiento__lte=fin_mes
    ).count()
    
    salidas_mes = MovimientoInventario.objects.filter(
        tipo_movimiento='SALIDA',
        estado='CONFIRMADO',
        fecha_movimiento__gte=inicio_mes,
        fecha_movimiento__lte=fin_mes
    ).count()
    
    # Paginación
    try:
        per_page = int(per_page)
        if per_page not in [10, 25, 50, 100]:
            per_page = 25
    except ValueError:
        per_page = 25
    
    paginator = Paginator(movimientos, per_page)
    page_number = request.GET.get('page', 1)
    movimientos_page = paginator.get_page(page_number)
    
    # Obtener listas para filtros
    bodegas = Bodega.objects.filter(activo=True).order_by('nombre')
    
    # Obtener nombre de bodega seleccionada si existe
    bodega_nombre = None
    if bodega_id:
        try:
            bodega_seleccionada = Bodega.objects.get(id=bodega_id)
            bodega_nombre = f"{bodega_seleccionada.codigo} - {bodega_seleccionada.nombre}"
        except Bodega.DoesNotExist:
            pass
    
    context = {
        'movimientos': movimientos_page,
        'page_obj': movimientos_page,
        'total_movimientos': total_movimientos,
        'movimientos_hoy': movimientos_hoy,
        'movimientos_pendientes': movimientos_pendientes,
        'ingresos_mes': ingresos_mes,
        'salidas_mes': salidas_mes,
        'bodegas': bodegas,
        'permisos': permisos,
        'search': search,
        'per_page': per_page,
        'sort_by': sort_by,
        'sort_order': sort_order,
        'tipo_movimiento': tipo_movimiento,
        'estado': estado,
        'bodega_id': bodega_id,
        'bodega_nombre': bodega_nombre,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'per_page': per_page,
        'active_menu': 'movimientos',
    }
    
    return render(request, 'inventario/movimientos_list.html', context)


@login_required
@require_http_methods(["GET", "POST"])
def movimiento_create(request):
    """Crear un nuevo movimiento de inventario"""
    
    # Verificar permisos
    permisos = {'crear': True}
    if hasattr(request.user, 'usuario_profile'):
        usuario = request.user.usuario_profile
        if usuario.rol and usuario.rol.permisos and isinstance(usuario.rol.permisos, dict):
            rol_permisos = usuario.rol.permisos.get('inventario', {})
            if rol_permisos:
                permisos = rol_permisos
    
    if not permisos.get('crear'):
        return JsonResponse({'success': False, 'message': 'No tienes permisos para crear movimientos'}, status=403)
    
    if request.method == 'GET':
        # Retornar datos para el formulario
        productos = Producto.objects.filter(estado='ACTIVO').values('id', 'sku', 'nombre', 'control_por_lote', 'control_por_serie')
        bodegas = Bodega.objects.filter(activo=True).values('id', 'codigo', 'nombre')
        proveedores = Proveedor.objects.filter(estado='ACTIVO').values('id', 'rut_nif', 'razon_social')
        unidades = UnidadMedida.objects.values('id', 'codigo', 'nombre')
        
        return JsonResponse({
            'success': True,
            'productos': list(productos),
            'bodegas': list(bodegas),
            'proveedores': list(proveedores),
            'unidades': list(unidades)
        })
    
    # POST - Crear movimiento
    try:
        data = json.loads(request.body)
        
        # Validaciones básicas
        tipo_movimiento = data.get('tipo_movimiento')
        fecha_movimiento_str = data.get('fecha_movimiento')
        producto_id = data.get('producto_id')
        cantidad = Decimal(str(data.get('cantidad', 0)))
        unidad_medida_id = data.get('unidad_medida_id')
        
        if not all([tipo_movimiento, fecha_movimiento_str, producto_id, cantidad, unidad_medida_id]):
            return JsonResponse({'success': False, 'message': 'Faltan campos obligatorios'}, status=400)
        
        # Convertir fecha string a datetime con timezone local
        fecha_naive = datetime.strptime(fecha_movimiento_str, '%Y-%m-%dT%H:%M')
        fecha_movimiento = timezone.make_aware(fecha_naive)
        
        producto = get_object_or_404(Producto, id=producto_id)
        unidad_medida = get_object_or_404(UnidadMedida, id=unidad_medida_id)
        
        # Validaciones según tipo de movimiento
        proveedor_id = data.get('proveedor_id')
        bodega_origen_id = data.get('bodega_origen_id')
        bodega_destino_id = data.get('bodega_destino_id')
        
        if tipo_movimiento in ['INGRESO', 'DEVOLUCION'] and not proveedor_id:
            return JsonResponse({'success': False, 'message': 'Proveedor es obligatorio para ingresos y devoluciones'}, status=400)
        
        if tipo_movimiento in ['SALIDA', 'AJUSTE', 'TRANSFERENCIA'] and not bodega_origen_id:
            return JsonResponse({'success': False, 'message': 'Bodega origen es obligatoria para este tipo de movimiento'}, status=400)
        
        if tipo_movimiento in ['INGRESO', 'TRANSFERENCIA'] and not bodega_destino_id:
            return JsonResponse({'success': False, 'message': 'Bodega destino es obligatoria para este tipo de movimiento'}, status=400)
        
        # Validar control por lote
        lote_id = data.get('lote_id')
        if producto.control_por_lote and not lote_id:
            return JsonResponse({'success': False, 'message': f'El producto {producto.sku} requiere control por lote'}, status=400)
        
        # Validar control por serie
        serie_value = data.get('serie')
        serie = serie_value.strip() if serie_value else ''
        if producto.control_por_serie and not serie:
            return JsonResponse({'success': False, 'message': f'El producto {producto.sku} requiere número de serie'}, status=400)
        
        # Obtener y limpiar campos opcionales
        doc_ref = data.get('documento_referencia')
        documento_referencia = doc_ref.strip() if doc_ref else ''
        
        motivo = data.get('motivo_ajuste')
        motivo_ajuste = motivo.strip() if motivo else ''
        
        obs = data.get('observaciones')
        observaciones = obs.strip() if obs else ''
        
        # Crear movimiento
        movimiento = MovimientoInventario.objects.create(
            tipo_movimiento=tipo_movimiento,
            fecha_movimiento=fecha_movimiento,
            producto=producto,
            proveedor_id=proveedor_id if proveedor_id else None,
            bodega_origen_id=bodega_origen_id if bodega_origen_id else None,
            bodega_destino_id=bodega_destino_id if bodega_destino_id else None,
            cantidad=cantidad,
            unidad_medida=unidad_medida,
            costo_unitario=Decimal(str(data.get('costo_unitario', 0))) if data.get('costo_unitario') else None,
            costo_total=Decimal(str(data.get('costo_total', 0))) if data.get('costo_total') else None,
            lote_id=lote_id if lote_id else None,
            serie=serie if serie else None,
            documento_padre_tipo=data.get('documento_padre_tipo'),
            documento_padre_id=data.get('documento_padre_id'),
            documento_referencia=documento_referencia,
            motivo_ajuste=motivo_ajuste,
            usuario=request.user.usuario_profile,
            observaciones=observaciones,
            estado=data.get('estado', 'PENDIENTE')
        )
        
        # Si el movimiento está confirmado, actualizar stock
        if movimiento.estado == 'CONFIRMADO':
            actualizar_stock_producto(movimiento)
        
        return JsonResponse({
            'success': True,
            'message': f'Movimiento {movimiento.get_tipo_movimiento_display()} creado exitosamente',
            'movimiento_id': movimiento.id
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Datos JSON inválidos'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
@require_http_methods(["GET", "POST"])
def movimiento_edit(request, movimiento_id):
    """Editar un movimiento de inventario"""
    
    movimiento = get_object_or_404(MovimientoInventario, id=movimiento_id)
    
    if request.method == 'GET':
        # GET es solo para ver detalles, no requiere validación de permisos de edición
        # Retornar datos del movimiento
        data = {
            'id': movimiento.id,
            'tipo_movimiento': movimiento.tipo_movimiento,
            'fecha_movimiento': movimiento.fecha_movimiento.strftime('%Y-%m-%d %H:%M'),
            'producto_id': movimiento.producto_id,
            'producto_sku': movimiento.producto.sku,
            'producto_nombre': movimiento.producto.nombre,
            'producto_texto': f"{movimiento.producto.sku} - {movimiento.producto.nombre}",
            'proveedor_id': movimiento.proveedor_id,
            'proveedor_texto': f"{movimiento.proveedor.rut_nif} - {movimiento.proveedor.razon_social}" if movimiento.proveedor else '',
            'bodega_origen_id': movimiento.bodega_origen_id,
            'bodega_origen_texto': f"{movimiento.bodega_origen.codigo} - {movimiento.bodega_origen.nombre}" if movimiento.bodega_origen else '',
            'bodega_destino_id': movimiento.bodega_destino_id,
            'bodega_destino_texto': f"{movimiento.bodega_destino.codigo} - {movimiento.bodega_destino.nombre}" if movimiento.bodega_destino else '',
            'cantidad': str(movimiento.cantidad),
            'unidad_medida_id': movimiento.unidad_medida_id,
            'costo_unitario': str(movimiento.costo_unitario) if movimiento.costo_unitario else '',
            'costo_total': str(movimiento.costo_total) if movimiento.costo_total else '',
            'lote_id': movimiento.lote_id,
            'lote_codigo': movimiento.lote.codigo_lote if movimiento.lote else '',
            'serie': movimiento.serie or '',
            'documento_padre_tipo': movimiento.documento_padre_tipo or '',
            'documento_padre_id': movimiento.documento_padre_id,
            'documento_referencia': movimiento.documento_referencia or '',
            'motivo_ajuste': movimiento.motivo_ajuste or '',
            'observaciones': movimiento.observaciones or '',
            'estado': movimiento.estado
        }
        
        return JsonResponse({'success': True, 'movimiento': data})
    
    # POST - Actualizar movimiento
    # Verificar permisos para editar
    permisos = {'editar': True}
    if hasattr(request.user, 'usuario_profile'):
        usuario = request.user.usuario_profile
        if usuario.rol and usuario.rol.permisos and isinstance(usuario.rol.permisos, dict):
            rol_permisos = usuario.rol.permisos.get('inventario', {})
            if rol_permisos:
                permisos = rol_permisos
    
    if not permisos.get('editar'):
        return JsonResponse({'success': False, 'message': 'No tienes permisos para editar movimientos'}, status=403)
    
    # No permitir editar movimientos confirmados
    if movimiento.estado == 'CONFIRMADO':
        return JsonResponse({'success': False, 'message': 'No se puede editar un movimiento confirmado'}, status=400)
    
    try:
        data = json.loads(request.body)
        
        # Actualizar campos permitidos
        movimiento.cantidad = Decimal(str(data.get('cantidad', movimiento.cantidad)))
        movimiento.costo_unitario = Decimal(str(data.get('costo_unitario'))) if data.get('costo_unitario') else None
        movimiento.costo_total = Decimal(str(data.get('costo_total'))) if data.get('costo_total') else None
        
        # Limpiar campos de texto opcionales
        doc_ref = data.get('documento_referencia')
        movimiento.documento_referencia = doc_ref.strip() if doc_ref else ''
        
        motivo = data.get('motivo_ajuste')
        movimiento.motivo_ajuste = motivo.strip() if motivo else ''
        
        obs = data.get('observaciones')
        movimiento.observaciones = obs.strip() if obs else ''
        
        # Verificar si cambió de PENDIENTE a CONFIRMADO
        estado_anterior = movimiento.estado
        movimiento.estado = data.get('estado', movimiento.estado)
        
        movimiento.save()
        
        # Si el movimiento pasó a CONFIRMADO, actualizar stock
        if estado_anterior == 'PENDIENTE' and movimiento.estado == 'CONFIRMADO':
            actualizar_stock_producto(movimiento)
        
        return JsonResponse({
            'success': True,
            'message': 'Movimiento actualizado exitosamente'
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Datos JSON inválidos'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def movimiento_delete(request, movimiento_id):
    """Anular un movimiento de inventario"""
    
    # Verificar permisos
    permisos = {'eliminar': True}
    if hasattr(request.user, 'usuario_profile'):
        usuario = request.user.usuario_profile
        if usuario.rol and usuario.rol.permisos and isinstance(usuario.rol.permisos, dict):
            rol_permisos = usuario.rol.permisos.get('inventario', {})
            if rol_permisos:
                permisos = rol_permisos
    
    if not permisos.get('eliminar'):
        return JsonResponse({'success': False, 'message': 'No tienes permisos para anular movimientos'}, status=403)
    
    movimiento = get_object_or_404(MovimientoInventario, id=movimiento_id)
    
    # No permitir anular movimientos confirmados (requiere un proceso especial)
    if movimiento.estado == 'CONFIRMADO':
        return JsonResponse({'success': False, 'message': 'No se puede anular un movimiento confirmado directamente. Contacte al administrador.'}, status=400)
    
    try:
        movimiento.estado = 'ANULADO'
        movimiento.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Movimiento anulado exitosamente'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
def exportar_movimientos_excel(request):
    """Exportar movimientos a Excel"""
    
    # Verificar permisos
    if hasattr(request.user, 'usuario_profile'):
        user_profile = request.user.usuario_profile
        if user_profile.rol and user_profile.rol.permisos and isinstance(user_profile.rol.permisos, dict):
            permisos = user_profile.rol.permisos.get('inventario', {})
            if not permisos.get('exportar', True):
                return JsonResponse({
                    'success': False,
                    'message': 'No tienes permisos para exportar movimientos'
                }, status=403)
    
    # Obtener los mismos filtros que la vista principal
    search = request.GET.get('search', '').strip()
    tipo_movimiento = request.GET.get('tipo_movimiento', '')
    estado = request.GET.get('estado', '')
    bodega_id = request.GET.get('bodega', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    
    # Query base
    movimientos = MovimientoInventario.objects.select_related(
        'producto', 'bodega_origen', 'bodega_destino', 'proveedor', 
        'usuario__user', 'lote', 'unidad_medida'
    ).all()
    
    # Aplicar los mismos filtros
    if search:
        movimientos = movimientos.filter(
            Q(producto__sku__icontains=search) |
            Q(producto__nombre__icontains=search) |
            Q(documento_referencia__icontains=search) |
            Q(serie__icontains=search) |
            Q(lote__codigo_lote__icontains=search)
        )
    
    if tipo_movimiento:
        movimientos = movimientos.filter(tipo_movimiento=tipo_movimiento)
    
    if estado:
        movimientos = movimientos.filter(estado=estado)
    
    if bodega_id:
        movimientos = movimientos.filter(
            Q(bodega_origen_id=bodega_id) | Q(bodega_destino_id=bodega_id)
        )
    
    if fecha_desde:
        movimientos = movimientos.filter(fecha_movimiento__gte=fecha_desde)
    
    if fecha_hasta:
        movimientos = movimientos.filter(fecha_movimiento__lte=fecha_hasta + ' 23:59:59')
    
    movimientos = movimientos.order_by('-fecha_movimiento')
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimientos Inventario"
    
    # Estilos
    header_fill = PatternFill(start_color="D20A11", end_color="D20A11", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = [
        'ID', 'Fecha Movimiento', 'Tipo', 'Estado', 'Producto SKU', 'Producto Nombre',
        'Cantidad', 'Unidad', 'Bodega Origen', 'Bodega Destino', 'Proveedor',
        'Lote', 'Serie', 'Costo Unitario', 'Costo Total', 'Doc. Referencia',
        'Motivo/Observaciones', 'Usuario', 'Fecha Registro'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Datos
    for row_num, mov in enumerate(movimientos, 2):
        ws.cell(row=row_num, column=1).value = mov.id
        ws.cell(row=row_num, column=2).value = mov.fecha_movimiento.strftime('%d/%m/%Y %H:%M')
        ws.cell(row=row_num, column=3).value = mov.get_tipo_movimiento_display()
        ws.cell(row=row_num, column=4).value = mov.get_estado_display()
        ws.cell(row=row_num, column=5).value = mov.producto.sku
        ws.cell(row=row_num, column=6).value = mov.producto.nombre
        ws.cell(row=row_num, column=7).value = float(mov.cantidad)
        ws.cell(row=row_num, column=8).value = mov.unidad_medida.codigo
        ws.cell(row=row_num, column=9).value = f"{mov.bodega_origen.codigo} - {mov.bodega_origen.nombre}" if mov.bodega_origen else ''
        ws.cell(row=row_num, column=10).value = f"{mov.bodega_destino.codigo} - {mov.bodega_destino.nombre}" if mov.bodega_destino else ''
        ws.cell(row=row_num, column=11).value = mov.proveedor.razon_social if mov.proveedor else ''
        ws.cell(row=row_num, column=12).value = mov.lote.codigo_lote if mov.lote else ''
        ws.cell(row=row_num, column=13).value = mov.serie or ''
        ws.cell(row=row_num, column=14).value = float(mov.costo_unitario) if mov.costo_unitario else None
        ws.cell(row=row_num, column=15).value = float(mov.costo_total) if mov.costo_total else None
        ws.cell(row=row_num, column=16).value = mov.documento_referencia or ''
        ws.cell(row=row_num, column=17).value = mov.motivo_ajuste or mov.observaciones or ''
        ws.cell(row=row_num, column=18).value = mov.usuario.user.username
        ws.cell(row=row_num, column=19).value = mov.created_at.strftime('%d/%m/%Y %H:%M')
        
        # Aplicar bordes
        for col_num in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col_num).border = border
    
    # Ajustar ancho de columnas
    for col_num in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 15
    
    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=movimientos_inventario_dulceria_lilis.xlsx'
    wb.save(response)
    
    return response


@login_required
@require_http_methods(["GET"])
def buscar_bodegas(request):
    """Buscar bodegas por nombre o código (para autocompletado)"""
    
    search = request.GET.get('q', '').strip()
    
    # Query base - solo bodegas activas
    bodegas = Bodega.objects.filter(activo=True)
    
    # Filtrar si hay búsqueda
    if search:
        bodegas = bodegas.filter(
            Q(nombre__icontains=search) |
            Q(codigo__icontains=search)
        )
    
    # Limitar a 20 resultados
    bodegas = bodegas.order_by('nombre')[:20]
    
    # Formatear resultados
    results = [
        {
            'id': bodega.id,
            'codigo': bodega.codigo,
            'nombre': bodega.nombre,
            'text': f"{bodega.codigo} - {bodega.nombre}"
        }
        for bodega in bodegas
    ]
    
    return JsonResponse({
        'success': True,
        'results': results
    })


@login_required
@require_http_methods(["GET"])
def buscar_productos(request):
    """Buscar productos por SKU o nombre (para autocompletado)"""
    
    search = request.GET.get('q', '').strip()
    
    # Query base - solo productos activos
    productos = Producto.objects.filter(estado='ACTIVO')
    
    # Filtrar si hay búsqueda
    if search:
        productos = productos.filter(
            Q(sku__icontains=search) |
            Q(nombre__icontains=search)
        )
    
    # Limitar a 20 resultados
    productos = productos.order_by('nombre')[:20]
    
    # Formatear resultados
    results = [
        {
            'id': producto.id,
            'sku': producto.sku,
            'nombre': producto.nombre,
            'text': f"{producto.sku} - {producto.nombre}",
            'control_por_lote': producto.control_por_lote,
            'control_por_serie': producto.control_por_serie
        }
        for producto in productos
    ]
    
    return JsonResponse({
        'success': True,
        'results': results
    })


@login_required
@require_http_methods(["GET"])
def buscar_proveedores(request):
    """Buscar proveedores por RUT o razón social (para autocompletado)"""
    
    search = request.GET.get('q', '').strip()
    
    # Query base - solo proveedores activos
    proveedores = Proveedor.objects.filter(estado='ACTIVO')
    
    # Filtrar si hay búsqueda
    if search:
        proveedores = proveedores.filter(
            Q(rut_nif__icontains=search) |
            Q(razon_social__icontains=search)
        )
    
    # Limitar a 20 resultados
    proveedores = proveedores.order_by('razon_social')[:20]
    
    # Formatear resultados
    results = [
        {
            'id': proveedor.id,
            'rut': proveedor.rut_nif,
            'razon_social': proveedor.razon_social,
            'text': f"{proveedor.rut_nif} - {proveedor.razon_social}"
        }
        for proveedor in proveedores
    ]
    
    return JsonResponse({
        'success': True,
        'results': results
    })
